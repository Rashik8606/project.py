from openpyxl import Workbook,load_workbook
import tkinter as tk
from tkinter import Label,Entry,Button

class data_entry(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title('data entry form')
        self.geometry('500x200')

        self.item_label = Label(self,text='Item')
        self.item_entry = Entry(self)

        self.qut_label = Label(self,text="quantity")
        self.qut_entry = Entry(self)

        self.prize_label = Label(self,text='Prize')
        self.prize_entry = Entry(self)

        self.add_btn = Button(self,text='add to excel',command=self.add_data)

        self.item_label.grid(row=0,column=0,padx=10,pady=10)
        self.item_entry.grid(row=0,column=1,padx=10,pady=10)

        self.qut_label.grid(row=1,column=0,padx=10,pady=10)
        self.qut_entry.grid(row=1,column=1,padx=10,pady=10)

        self.prize_label.grid(row=2,column=0,padx=10,pady=10)
        self.prize_entry.grid(row=2,column=1,padx=10,pady=10)

        self.add_btn.grid(row=3,column=3,padx=10,pady=10)

    def add_data(self):
        item = self.item_entry.get()
        qut = self.qut_entry.get()
        prize = self.prize_entry.get()

        if item and qut and prize:
            self.write_to_excel(item,qut,prize)

        self.item_entry.delete(0,'end')
        self.qut_entry.delete(0,'end')
        self.prize_entry.delete(0,'end')

    def write_to_excel(self,item,qut,prize):
        try:
            workbook=load_workbook('item.xlsx')
        except FileNotFoundError:
            workbook = Workbook()

        sheet = workbook.active
        row = [item,qut,prize]
        sheet.append(row)

        workbook.save('item.xlsx')

app=data_entry()
app.mainloop()
